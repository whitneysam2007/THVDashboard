import json
from pathlib import Path

from openpyxl import load_workbook

workbook_path = Path('/home/ubuntu/upload/Topdonors,livelist_HumansthatkeepHVAliveFall2025-Present.xlsx')
target_sheet = 'Donors $500-$4,999'

workbook = load_workbook(workbook_path, read_only=True, data_only=True)
if target_sheet not in workbook.sheetnames:
    raise RuntimeError(f'Expected worksheet {target_sheet!r} was not found. Available sheet names: {workbook.sheetnames!r}')

sheet = workbook[target_sheet]
rows = list(sheet.iter_rows(values_only=True))
header = [str(value).strip() if value is not None else '' for value in (rows[0] if rows else [])]
sample_rows = [list(row) for row in rows[1:6]]
nonempty_rows = [list(row) for row in rows[1:] if any(value not in (None, '') for value in row)]

print(json.dumps({
    'targetWorksheet': target_sheet,
    'rowCount': len(nonempty_rows),
    'columnCount': len(header),
    'headers': header,
    'sampleRows': sample_rows,
}, default=str, indent=2))
