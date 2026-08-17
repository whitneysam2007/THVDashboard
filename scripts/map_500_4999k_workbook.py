import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

workbook_path = Path('/home/ubuntu/upload/Topdonors,livelist_HumansthatkeepHVAliveFall2025-Present.xlsx')
sheet_name = 'Donors $500-$4,999'
workbook = load_workbook(workbook_path, read_only=True, data_only=True)
sheet = workbook[sheet_name]
rows = list(sheet.iter_rows(values_only=True))

def date_value(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    return str(value).strip() if value not in (None, '') else None

records = []
for index, row in enumerate(rows[1:], start=2):
    if not any(value not in (None, '') for value in row):
        continue
    record = {
        'worksheetRow': index,
        'donationDate': date_value(row[0]),
        'amount': row[1],
        'givingType': str(row[2]).strip() if row[2] else None,
        'firstName': str(row[3]).strip() if row[3] else None,
        'lastName': str(row[4]).strip() if row[4] else None,
        'email': str(row[5]).strip() if row[5] else None,
        'address': str(row[6]).strip() if row[6] else None,
        'cityState': str(row[7]).strip() if row[7] else None,
        'zip': str(int(row[8])) if isinstance(row[8], float) and row[8].is_integer() else str(row[8]).strip() if row[8] else None,
        'annualAmount': row[9],
        'initialContact': str(row[10]).strip() if row[10] else None,
        'thankYouSent': str(row[11]).strip() if row[11] else None,
        'thankYouDate': date_value(row[12]),
        'connection': str(row[13]).strip() if row[13] else None,
        'paymentMethod': str(row[14]).strip() if row[14] else None,
        'notes': str(row[15]).strip() if row[15] else None,
    }
    record['suggestedPortfolio'] = 'monthly-giving' if (record['givingType'] or '').lower() == 'monthly' else 'donors-500-5k'
    records.append(record)

eligible_records = [
    record for record in records
    if isinstance(record['amount'], (int, float))
    and isinstance(record['annualAmount'], (int, float))
    and 500 <= record['annualAmount'] < 5000
    and (record['givingType'] or '').strip().lower() != 'monthly'
    and record['firstName']
]
excluded_records = [record for record in records if record not in eligible_records]
report = {
    'targetWorksheet': sheet_name,
    'eligibleDonors500To4999': eligible_records,
    'excludedRows': [{
        'worksheetRow': record['worksheetRow'],
        'name': ' '.join(part for part in [record['firstName'], record['lastName']] if part),
        'givingType': record['givingType'],
        'annualAmount': record['annualAmount'],
    } for record in excluded_records],
}
Path('/home/ubuntu/thv-donor-dashboard/docs/donors-500-4999-import-map.json').write_text(json.dumps(report, indent=2))
print(json.dumps({
    'targetWorksheet': sheet_name,
    'eligibleRecordCount': len(eligible_records),
    'excludedRowCount': len(excluded_records),
    'reportPath': '/home/ubuntu/thv-donor-dashboard/docs/donors-500-4999-import-map.json',
}, indent=2))
